"""Excel spreadsheet (.xlsx) extraction via openpyxl.

Requires the optional ``openpyxl`` package:
    pip install 'embd[xlsx]'

Each worksheet becomes one PageText (page_num = sheet index, 1-based).
Cell values in each row are joined with `` | `` so the chunker and LLM
see a readable flat representation rather than raw CSV.
"""
from __future__ import annotations

import logging
from pathlib import Path

from ..config import IngestionConfig
from .extractor import PageText

logger = logging.getLogger(__name__)


def extract_xlsx_pages(path: Path, _ingestion: IngestionConfig) -> list[PageText]:
    """Extract text from each sheet of an .xlsx file."""
    try:
        from openpyxl import load_workbook  # type: ignore[import-untyped]
    except ImportError:
        logger.error(
            "openpyxl is required for .xlsx files. "
            "Install with: pip install 'embd[xlsx]'"
        )
        return []

    try:
        wb = load_workbook(str(path), read_only=True, data_only=True)
    except Exception as exc:
        logger.warning("Cannot read '%s': %s — skipping file.", path.name, exc)
        return []

    pages: list[PageText] = []
    for i, ws in enumerate(wb.worksheets, start=1):
        rows: list[str] = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) for c in row if c is not None]
            if cells:
                rows.append(" | ".join(cells))
        text = "\n".join(rows).strip()
        if text:
            pages.append(PageText(page_num=i, text=text))

    wb.close()

    if not pages:
        logger.warning("'%s' yielded no extractable text.", path.name)

    return pages
